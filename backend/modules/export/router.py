import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

router = APIRouter(prefix="/api/export", tags=["export"])


class ExportItem(BaseModel):
    name: str
    category: str
    price: float
    commission_rate: float
    score: float
    recommend: bool
    summary: str
    hook: str
    script: str
    caption: str
    hashtags: str
    affiliate_url: str
    role: str


class ExportRequest(BaseModel):
    basket_name: Optional[str] = "ตะกร้าสินค้า Affiliate"
    items: List[ExportItem]


@router.post("/excel")
def export_excel(body: ExportRequest):
    wb = Workbook()
    ws = wb.active
    ws.title = "ตะกร้าสินค้า"

    # สี
    orange = "FF6B00"
    light_orange = "FFF3E0"
    green = "2E7D32"
    light_green = "E8F5E9"
    gray = "F5F5F5"
    white = "FFFFFF"

    # Header row
    headers = [
        "ลำดับ", "ชื่อสินค้า", "หมวดหมู่", "ราคา (฿)",
        "คอมมิชชัน %", "คะแนนน่าทำ", "บทบาทในตะกร้า",
        "น่าทำเพราะ...",
        "🎬 Hook เปิดคลิป",
        "📝 Script พูดในคลิป",
        "✍️ Caption โพสต์",
        "# Hashtag",
        "🔗 ลิงก์ Affiliate"
    ]

    col_widths = [6, 35, 15, 10, 12, 12, 15, 45, 45, 55, 55, 35, 55]

    # Title row
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"🛒 {body.basket_name} — สร้างโดย Shopee Affiliate Basket Manager"
    title_cell.font = Font(bold=True, size=13, color=white)
    title_cell.fill = PatternFill("solid", fgColor=orange)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, size=10, color=white)
        cell.fill = PatternFill("solid", fgColor="BF360C")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 32

    # Data rows
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, item in enumerate(body.items):
        row = i + 3
        is_recommend = item.recommend
        bg = light_green if is_recommend else light_orange
        emoji = "✅" if is_recommend else "⚠️"

        values = [
            i + 1,
            item.name,
            item.category,
            item.price,
            f"{item.commission_rate}%",
            item.score,
            item.role,
            f"{emoji} {item.summary}",
            item.hook,
            item.script,
            item.caption,
            item.hashtags,
            "🛍 เปิดลิงก์ Shopee",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg if col > 1 else (
                "E8F5E9" if is_recommend else "FFF3E0"))
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal=(
                "center" if col in [1, 4, 5, 6] else "left"))
            cell.border = border
            cell.font = Font(size=9)

            # ไฮไลต์คอลัมน์คะแนน
            if col == 6:
                color = "1B5E20" if item.score >= 70 else ("E65100" if item.score >= 55 else "757575")
                cell.font = Font(size=11, bold=True, color=color)

            # ลิงก์ Affiliate — แสดงเป็น Hyperlink คลิกได้ ไม่โชว์ URL ยาว
            if col == 13:
                cell.hyperlink = item.affiliate_url
                cell.font = Font(size=9, color="1565C0", underline="single")
                cell.alignment = Alignment(vertical="top", horizontal="center")

        ws.row_dimensions[row].height = 80

    # Freeze หัวตาราง
    ws.freeze_panes = "A3"

    # บันทึก
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"affiliate_basket.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
